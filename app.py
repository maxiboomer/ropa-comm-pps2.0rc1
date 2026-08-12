#!/usr/bin/env python3
"""
RoPA — Web App
LGPD compliance
"""

import csv
import io
import json
import os
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path

from authlib.integrations.flask_client import OAuth
from flask import (Flask, Response, flash, redirect, render_template,
                   request, session, url_for)

import cnil_pia_importer
from modelo_ppsi import (
    CAMPOS_VALIDACAO, SITUACOES, CATEGORIAS_DADOS_FCI, JSON_FIELDS,
    migrar_schema, proxima_versao, campos_estruturais, preenchido,
    parse_lista, parse_dict_tipos, parse_estimativa, parse_json,
    lista_para_texto, dict_tipos_para_texto, dict_estimativa_para_texto,
    calcular_risco, SITUACOES_RIPD, PRINCIPIOS_LGPD, DIREITOS_TITULARES,
    CRITERIO_GERAL_LABELS, CRITERIO_ESPECIFICO_LABELS,
    # Porte do repo `ropa` — matriz 5×5, gatilhos, sugestões, aprovações
    nivel_risco, consolidar_risco, NIVEL_LABEL, NIVEL_COR, PROB_LABELS, IMPACTO_LABELS,
    CATEGORIAS_RISCO, RISCOS_TIPICOS_POR_FATOR, SALVAGUARDAS_TIPICAS, TIPO_SALVAGUARDA_LABEL,
    gatilhos_ripd, _fatores_atividade, sugerir_riscos, PAPEIS_APROVACAO, papel_label,
)

# ── ropa.py shared logic ──────────────────────────────────────────────────────
DB_PATH = Path(os.environ.get("ROPA_DB_PATH", Path(__file__).parent / "ropa.db"))
EXPORT_DIR = Path(os.environ.get("ROPA_DATA_DIR", Path(__file__).parent)) / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

# Identidade institucional — configurável via env (genérico por padrão)
ORGANIZACAO   = os.environ.get("ROPA_ORGANIZACAO", "Organização")
UNIDADE       = os.environ.get("ROPA_UNIDADE", "Unidade de Proteção de Dados")
ENCARREGADO   = os.environ.get("ROPA_ENCARREGADO", "Encarregado(a) de Proteção de Dados")
NORMAS_RODAPE = os.environ.get(
    "ROPA_NORMAS_REFERENCIA",
    "Documento produzido nos termos da LGPD – Lei 13.709/2018, Art. 37",
)

BASES_LEGAIS = {
    "I":    "Consentimento do titular (Art. 7º, I)",
    "II":   "Obrigação legal ou regulatória (Art. 7º, II)",
    "III":  "Execução de políticas públicas (Art. 7º, III)",
    "IV":   "Estudos por órgão de pesquisa (Art. 7º, IV)",
    "V":    "Execução de contrato (Art. 7º, V)",
    "VI":   "Exercício regular de direitos (Art. 7º, VI)",
    "VII":  "Proteção da vida ou incolumidade física (Art. 7º, VII)",
    "VIII": "Tutela da saúde (Art. 7º, VIII)",
    "IX":   "Legítimo interesse (Art. 7º, IX)",
    "X":    "Proteção do crédito (Art. 7º, X)",
    "S-I":  "Dados sensíveis – Consentimento específico (Art. 11, I)",
    "S-II": "Dados sensíveis – Obrigação legal / exercício de direitos / políticas públicas (Art. 11, II)",
}

CAMPOS_VALIDACAO = CAMPOS_VALIDACAO  # vindo de modelo_ppsi.py (soma=100, alinhado ao Guia PPSI 2.0)

EXEMPLOS = [
    dict(
        nome_atividade="Folha de Pagamento de Colaboradores",
        finalidade="Processamento da remuneração, encargos e benefícios dos colaboradores da Organização",
        base_legal="II",
        categorias_titulares="Colaboradores efetivos, comissionados e contratados",
        categorias_dados="Nome, CPF, matrícula, conta bancária, dependentes, dados previdenciários",
        dados_sensiveis=0,
        destinatarios="Órgão de arrecadação federal, instituição financeira pagadora, sistema de gestão de pessoas",
        transferencia_inter="N/A",
        prazo_retencao="20 anos para documentos trabalhistas; dados previdenciários: permanente",
        medidas_seguranca="Acesso restrito por perfil na área de gestão de pessoas; autenticação forte; canais cifrados com órgãos externos",
        unidade_controladora="Unidade de Gestão de Pessoas",
        sistema_sei="",
        observacoes="",
    ),
    dict(
        nome_atividade="Recrutamento e Seleção de Pessoal",
        finalidade="Gestão de processos seletivos e concursos para ingresso de novos colaboradores",
        base_legal="III",
        categorias_titulares="Candidatos inscritos em processos seletivos",
        categorias_dados="Nome, CPF, e-mail, telefone, escolaridade, histórico profissional, fotografia",
        dados_sensiveis=0,
        destinatarios="Comissão organizadora do certame; sistema oficial de inscrições",
        transferencia_inter="N/A",
        prazo_retencao="5 anos após o encerramento do certame, conforme normas internas",
        medidas_seguranca="Acesso restrito à comissão; ambiente controlado de provas; registros de auditoria",
        unidade_controladora="Unidade de Gestão de Pessoas",
        sistema_sei="",
        observacoes="",
    ),
    dict(
        nome_atividade="Atendimento ao Cidadão (Protocolo e Ouvidoria)",
        finalidade="Registro e tramitação de manifestações, solicitações e pedidos de informação dos cidadãos",
        base_legal="III",
        categorias_titulares="Cidadãos, solicitantes, manifestantes",
        categorias_dados="Nome, CPF, e-mail, telefone, conteúdo da manifestação, dados do atendimento",
        dados_sensiveis=0,
        destinatarios="Áreas internas responsáveis pela resposta; ouvidoria",
        transferencia_inter="N/A",
        prazo_retencao="Prazo legal de guarda de documentos administrativos, conforme normas internas",
        medidas_seguranca="Sistema de protocolo com perfis de acesso; sigilo das manifestações; trilha de auditoria",
        unidade_controladora="Unidade de Atendimento ao Cidadão",
        sistema_sei="",
        observacoes="",
    ),
    dict(
        nome_atividade="Controle de Acesso às Dependências",
        finalidade="Segurança patrimonial e controle de acesso de colaboradores e visitantes às dependências",
        base_legal="IX",
        categorias_titulares="Colaboradores, visitantes, prestadores de serviço",
        categorias_dados="Nome, CPF, horários de entrada e saída, imagem de identificação, dados biométricos",
        dados_sensiveis=1,
        destinatarios="Equipe de segurança institucional; autoridades policiais (incidentes); sem compartilhamento rotineiro",
        transferencia_inter="N/A",
        prazo_retencao="30 dias em sobrescrita contínua; incidentes: até encerramento de apuração",
        medidas_seguranca="Sistema de acesso com registro individual; sala de monitoramento com controle de acesso; sem transmissão externa",
        unidade_controladora="Unidade de Infraestrutura e Segurança",
        sistema_sei="",
        observacoes="RIPD recomendado (Art. 10, §3 LGPD – legítimo interesse + Art. 5º, II – dado sensível biométrico)",
    ),
    dict(
        nome_atividade="Processo Administrativo Disciplinar",
        finalidade="Apuração de irregularidades funcionais de colaboradores, conforme legislação aplicável",
        base_legal="VI",
        categorias_titulares="Colaboradores investigados, testemunhas, denunciantes",
        categorias_dados="Nome, matrícula, histórico funcional, depoimentos, documentos sigilosos",
        dados_sensiveis=0,
        destinatarios="Comissão processante, autoridade máxima, órgão de controle (se cabível)",
        transferencia_inter="N/A",
        prazo_retencao="10 anos após arquivamento; condenações: permanente",
        medidas_seguranca="Processo eletrônico com restrição de acesso; perfis específicos; impressão controlada",
        unidade_controladora="Unidade de Administração",
        sistema_sei="",
        observacoes="Dados de caráter sigiloso – acesso restrito nos termos da legislação de acesso à informação",
    ),
]


# ── DB helpers ────────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_conn() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS atividades (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_atividade        TEXT NOT NULL,
            finalidade            TEXT,
            base_legal            TEXT,
            categorias_titulares  TEXT,
            categorias_dados      TEXT,
            dados_sensiveis       INTEGER DEFAULT 0,
            destinatarios         TEXT,
            transferencia_inter   TEXT,
            prazo_retencao        TEXT,
            medidas_seguranca     TEXT,
            unidade_controladora  TEXT,
            sistema_sei           TEXT,
            observacoes           TEXT,
            criado_em             TEXT DEFAULT (datetime('now','localtime')),
            atualizado_em         TEXT DEFAULT (datetime('now','localtime')),
            ativo                 INTEGER DEFAULT 1
        )""")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS historico (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            atividade_id  INTEGER,
            campo         TEXT,
            valor_antigo  TEXT,
            valor_novo    TEXT,
            alterado_em   TEXT DEFAULT (datetime('now','localtime'))
        )""")
        migrar_schema(conn)


def todos_registros() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM atividades WHERE ativo=1 ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


def get_atividade(atividade_id: int):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM atividades WHERE id=? AND ativo=1", (atividade_id,)
        ).fetchone()
    return dict(row) if row else None


def pontuacao(atividade: dict) -> tuple[int, list]:
    score = 0
    faltando = []
    for campo, (descricao, peso) in CAMPOS_VALIDACAO.items():
        if preenchido(campo, atividade.get(campo)):
            score += peso
        else:
            faltando.append((descricao, peso))
    return score, faltando


def score_class(score: int) -> str:
    if score >= 80:
        return "success"
    if score >= 50:
        return "warning"
    return "danger"


# ── Flask app ─────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "ropa-dev-only-change-in-prod")

# Hardening de segurança (top-4 fix)
app.config.update(
    SESSION_COOKIE_SECURE=True,   # só envia cookie sobre HTTPS
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",  # mitigação CSRF
    MAX_CONTENT_LENGTH=8 * 1024 * 1024,  # limite upload
)

# ── Keycloak OIDC ─────────────────────────────────────────────────────────────

KEYCLOAK_MOCK_MODE = os.environ.get("KEYCLOAK_MOCK", "").lower() in ("1", "true", "yes")

if KEYCLOAK_MOCK_MODE:
    # Mock: endpoints rodam na mesma porta via blueprint
    from keycloak_blueprint import register_mock_keycloak
    ROPA_BASE_URL = os.environ.get("ROPA_BASE_URL", "http://127.0.0.1:8000")
    register_mock_keycloak(app, base_url=ROPA_BASE_URL)
    KEYCLOAK_URL = f"{ROPA_BASE_URL}/mock-kc"
else:
    KEYCLOAK_URL = os.environ.get("KEYCLOAK_URL", "http://localhost:8080")

KEYCLOAK_REALM = os.environ.get("KEYCLOAK_REALM", "ropa")
KEYCLOAK_CLIENT_ID = os.environ.get("KEYCLOAK_CLIENT_ID", "ropa-web")
KEYCLOAK_CLIENT_SECRET = os.environ.get("KEYCLOAK_CLIENT_SECRET", "")

OIDC_BASE = f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}"
OIDC_DISCOVERY = f"{OIDC_BASE}/.well-known/openid-configuration"

oauth = OAuth(app)
oauth.register(
    name="keycloak",
    client_id=KEYCLOAK_CLIENT_ID,
    client_secret=KEYCLOAK_CLIENT_SECRET,
    server_metadata_url=OIDC_DISCOVERY,
    client_kwargs={"scope": "openid email profile"},
)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated


app.jinja_env.globals.update(
    BASES_LEGAIS=BASES_LEGAIS,
    CAMPOS_VALIDACAO=CAMPOS_VALIDACAO,
    SITUACOES=SITUACOES,
    CATEGORIAS_DADOS_FCI=CATEGORIAS_DADOS_FCI,
    ORGANIZACAO=ORGANIZACAO,
    UNIDADE=UNIDADE,
    ENCARREGADO=ENCARREGADO,
    NORMAS_RODAPE=NORMAS_RODAPE,
    pontuacao=pontuacao,
    score_class=score_class,
    lista_para_texto=lista_para_texto,
    dict_tipos_para_texto=dict_tipos_para_texto,
    dict_estimativa_para_texto=dict_estimativa_para_texto,
    parse_json=parse_json,
    calcular_risco=calcular_risco,
    SITUACOES_RIPD=SITUACOES_RIPD,
    PRINCIPIOS_LGPD=PRINCIPIOS_LGPD,
    DIREITOS_TITULARES=DIREITOS_TITULARES,
    CRITERIO_GERAL_LABELS=CRITERIO_GERAL_LABELS,
    CRITERIO_ESPECIFICO_LABELS=CRITERIO_ESPECIFICO_LABELS,
    # Porte do repo `ropa` — matriz 5×5, gatilhos, sugestões, aprovações
    nivel_risco=nivel_risco, consolidar_risco=consolidar_risco,
    NIVEL_LABEL=NIVEL_LABEL, NIVEL_COR=NIVEL_COR,
    PROB_LABELS=PROB_LABELS, IMPACTO_LABELS=IMPACTO_LABELS,
    CATEGORIAS_RISCO=CATEGORIAS_RISCO,
    RISCOS_TIPICOS_POR_FATOR=RISCOS_TIPICOS_POR_FATOR,
    SALVAGUARDAS_TIPICAS=SALVAGUARDAS_TIPICAS, TIPO_SALVAGUARDA_LABEL=TIPO_SALVAGUARDA_LABEL,
    gatilhos_ripd=gatilhos_ripd, sugerir_riscos=sugerir_riscos,
    PAPEIS_APROVACAO=PAPEIS_APROVACAO, papel_label=papel_label,
    now=datetime.now,
)


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login")
def login():
    if "user" in session:
        return redirect(url_for("index"))
    error = request.args.get("error")
    return render_template("login.html", error=error)


@app.route("/login/keycloak")
def login_keycloak():
    redirect_uri = url_for("auth_callback", _external=True)
    if KEYCLOAK_MOCK_MODE:
        # Constrói a URL de autorização no host do próprio request (não no
        # ISSUER cacheado), para que o browser consiga alcançar o mock-kc
        # tanto via 127.0.0.1 quanto via URL pública (Cloudflare Tunnel).
        import secrets as _secrets
        import urllib.parse as _up
        state = _secrets.token_urlsafe(16)
        nonce = _secrets.token_urlsafe(16)
        session["oauth_state"] = state
        auth_url = f"{request.host_url.rstrip('/')}/mock-kc/realms/{KEYCLOAK_REALM}/protocol/openid-connect/auth"
        params = _up.urlencode({
            "response_type": "code",
            "client_id": KEYCLOAK_CLIENT_ID,
            "redirect_uri": redirect_uri,
            "scope": "openid email profile",
            "state": state,
            "nonce": nonce,
        })
        return redirect(f"{auth_url}?{params}")
    try:
        return oauth.keycloak.authorize_redirect(redirect_uri)
    except Exception:
        return redirect(url_for("login", error="Servidor Keycloak indisponível. Contate a Unidade de TI."))


@app.route("/auth/callback")
def auth_callback():
    import requests as req

    if KEYCLOAK_MOCK_MODE:
        # Mock: trocar o code manualmente (sem verificação JWT)
        code = request.args.get("code", "")
        if not code:
            return redirect(url_for("login", error="Falha na autenticação."))
        try:
            # Em modo mock, chamar os endpoints localmente para evitar
            # deadlock (request interno que sai pela URL pública e volta pelo
            # Caddy para o próprio gunicorn ocupado).
            _mock_base = os.environ.get("ROPA_MOCK_INTERNAL_URL", "http://127.0.0.1:5000")
            resp = req.post(f"{_mock_base}/mock-kc/realms/{KEYCLOAK_REALM}/protocol/openid-connect/token", data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": os.environ.get("ROPA_BASE_URL", "http://localhost:5000") + "/auth/callback",
                "client_id": KEYCLOAK_CLIENT_ID,
                "client_secret": KEYCLOAK_CLIENT_SECRET,
            })
            token = resp.json()
            if "error" in token:
                return redirect(url_for("login", error="Código expirado. Tente novamente."))
            ui_resp = req.get(f"{_mock_base}/mock-kc/realms/{KEYCLOAK_REALM}/protocol/openid-connect/userinfo",
                              headers={"Authorization": f"Bearer {token['access_token']}"})
            userinfo = ui_resp.json()
        except Exception:
            return redirect(url_for("login", error="Falha na autenticação."))
    else:
        # Produção: fluxo OIDC padrão via authlib
        try:
            token = oauth.keycloak.authorize_access_token()
            userinfo = token.get("userinfo", {})
            if not userinfo:
                userinfo = oauth.keycloak.userinfo()
        except Exception:
            return redirect(url_for("login", error="Falha na autenticação. Tente novamente."))

    session["user"] = {
        "sub": userinfo.get("sub", ""),
        "name": userinfo.get("name", userinfo.get("preferred_username", "Usuário")),
        "email": userinfo.get("email", ""),
        "username": userinfo.get("preferred_username", ""),
        "roles": userinfo.get("realm_access", {}).get("roles", []),
    }
    session["id_token"] = token.get("id_token", "")

    flash(f"Bem-vindo(a), {session['user']['name']}!", "success")
    next_url = request.args.get("next", url_for("index"))
    return redirect(next_url)


@app.route("/logout")
def logout():
    id_token = session.pop("id_token", "")
    session.pop("user", None)
    session.clear()

    post_logout_uri = url_for("login", _external=True)
    keycloak_logout = (
        f"{OIDC_BASE}/protocol/openid-connect/logout"
        f"?id_token_hint={id_token}"
        f"&post_logout_redirect_uri={post_logout_uri}"
    )
    return redirect(keycloak_logout)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    registros = todos_registros()
    scores = [pontuacao(r)[0] for r in registros]
    media = sum(scores) / len(scores) if scores else 0
    completos = sum(1 for s in scores if s >= 80)
    sensiveis = sum(1 for r in registros if r.get("dados_sensiveis"))

    # Ripd status por atividade
    ripd_status = {}
    with get_conn() as conn:
        for row in conn.execute("SELECT atividade_id, situacao FROM ripds"):
            ripd_status[row["atividade_id"]] = row["situacao"]

    # Distribuição por unidade (completude média) e base legal
    por_unidade = {}
    por_base = {}
    for r in registros:
        sc = pontuacao(r)[0]
        u = r.get("unidade_controladora") or "Sem unidade"
        por_unidade.setdefault(u, []).append(sc)
        bl = r.get("base_legal") or "—"
        por_base[bl] = por_base.get(bl, 0) + 1

    chart_unidades = [{"label": k, "media": round(sum(v) / len(v))} for k, v in por_unidade.items()]
    chart_bases = [{"label": k, "qtd": v} for k, v in por_base.items()]

    # Alertas de ação
    alertas = []
    abaixo = [r for r in registros if pontuacao(r)[0] < 80]
    if abaixo:
        alertas.append({"tipo": "danger", "icone": "bi-exclamation-triangle",
                        "titulo": f"{len(abaixo)} registro(s) abaixo de 80% de completude",
                        "sub": "Priorize o preenchimento destes registros.", "url": url_for("validar")})
    ripd_recomendados = []
    ripd_desatualizados = []
    for r in registros:
        rc = calcular_risco(r)
        st = ripd_status.get(r["id"])
        if rc["recomenda"] and not st:
            ripd_recomendados.append(r)
        if st == "desatualizado":
            ripd_desatualizados.append(r)
    if ripd_desatualizados:
        alertas.append({"tipo": "danger", "icone": "bi-arrow-repeat",
                        "titulo": f"{len(ripd_desatualizados)} RIPD(s) desatualizado(s) — revisar",
                        "sub": "A atividade vinculada mudou; o RIPD precisa de revisão.", "url": url_for("listar")})
    if ripd_recomendados:
        alertas.append({"tipo": "warn", "icone": "bi-file-earmark-text",
                        "titulo": f"{len(ripd_recomendados)} operação(ões) com RIPD recomendado",
                        "sub": "Indício de alto risco — elabore o RIPD (controle 23.3).", "url": url_for("listar")})

    return render_template(
        "index.html",
        registros=registros, scores=scores, media=media, completos=completos, sensiveis=sensiveis,
        chart_unidades=chart_unidades, chart_bases=chart_bases, alertas=alertas, ripd_status=ripd_status,
    )


@app.route("/atividades")
@login_required
def listar():
    registros = todos_registros()
    ripd_status = {}
    with get_conn() as conn:
        for row in conn.execute("SELECT atividade_id, situacao, versao FROM ripds"):
            ripd_status[row["atividade_id"]] = {"situacao": row["situacao"], "versao": row["versao"]}
    return render_template("listar.html", registros=registros, ripd_status=ripd_status)


@app.route("/atividades/nova", methods=["GET", "POST"])
@login_required
def nova():
    if request.method == "POST":
        dados = _form_to_dict(request.form)
        dados["versao"] = "1.0"
        dados["situacao"] = dados.get("situacao") or "em_andamento"
        cols = _COLUNAS_INSERT
        sql = (f"INSERT INTO atividades ({','.join(cols)}) "
               f"VALUES ({','.join(':' + c for c in cols)})")
        with get_conn() as conn:
            cur = conn.execute(sql, dados)
            novo_id = cur.lastrowid
            _registrar_versao(conn, novo_id, "1.0", "v1.0 – criação do registro",
                              _snapshot_json(dados), _responsavel_atual())
        flash(f"Atividade #{novo_id} criada com sucesso (versão 1.0).", "success")
        return redirect(url_for("ver", atividade_id=novo_id))
    return render_template("form.html", atividade=None, titulo="Nova Atividade")


@app.route("/atividades/<int:atividade_id>")
@login_required
def ver(atividade_id):
    atividade = get_atividade(atividade_id)
    if not atividade:
        flash(f"Atividade #{atividade_id} não encontrada.", "danger")
        return redirect(url_for("listar"))
    score, faltando = pontuacao(atividade)
    historico = []
    versoes = []
    risco = calcular_risco(atividade)
    ripd = None
    with get_conn() as conn:
        historico = conn.execute(
            "SELECT * FROM historico WHERE atividade_id=? ORDER BY alterado_em DESC LIMIT 20",
            (atividade_id,)
        ).fetchall()
        versoes = conn.execute(
            "SELECT * FROM versoes WHERE atividade_id=? ORDER BY id DESC LIMIT 30",
            (atividade_id,)
        ).fetchall()
        ripd_row = conn.execute(
            "SELECT * FROM ripds WHERE atividade_id=? ORDER BY id DESC LIMIT 1",
            (atividade_id,)
        ).fetchone()
        if ripd_row:
            ripd = dict(ripd_row)
    return render_template(
        "ver.html",
        atividade=atividade,
        score=score,
        faltando=faltando,
        historico=historico,
        versoes=versoes,
        risco=risco,
        ripd=ripd,
        base_desc=BASES_LEGAIS.get(atividade.get("base_legal", ""), "—"),
    )


@app.route("/atividades/<int:atividade_id>/editar", methods=["GET", "POST"])
@login_required
def editar(atividade_id):
    atividade = get_atividade(atividade_id)
    if not atividade:
        flash(f"Atividade #{atividade_id} não encontrada.", "danger")
        return redirect(url_for("listar"))

    if request.method == "POST":
        novos = _form_to_dict(request.form)
        with get_conn() as conn:
            for campo in CAMPOS_VALIDACAO:
                if novos.get(campo) != atividade.get(campo):
                    conn.execute("""
                        INSERT INTO historico (atividade_id, campo, valor_antigo, valor_novo)
                        VALUES (?,?,?,?)
                    """, (atividade_id, campo, atividade.get(campo), novos.get(campo)))

            # Versionamento semântico (Guia 4.1.6)
            estrutural = any(novos.get(c) != atividade.get(c) for c in campos_estruturais())
            nova_versao = proxima_versao(atividade.get("versao"), estrutural)
            sintese = f"v{nova_versao} – " + _sintese_alteracoes(atividade, novos)
            novos["versao"] = nova_versao
            novos["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            novos["id"] = atividade_id
            _registrar_versao(conn, atividade_id, nova_versao, sintese,
                              _snapshot_json({**atividade, **novos}), _responsavel_atual())

            set_clause = ",".join(f"{c}=:{c}" for c in _COLUNAS_INSERT)
            conn.execute(
                f"UPDATE atividades SET {set_clause}, atualizado_em=:atualizado_em WHERE id=:id",
                novos,
            )
            # P5: alteração estrutural no ROPA marca RIPDs vinculados como "revisar"
            if estrutural:
                conn.execute(
                    "UPDATE ripds SET situacao='desatualizado', atualizado_em=? "
                    "WHERE atividade_id=? AND situacao NOT IN ('desatualizado')",
                    (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), atividade_id),
                )
        flash(f"Atividade #{atividade_id} atualizada (versão {nova_versao}).", "success")
        return redirect(url_for("ver", atividade_id=atividade_id))

    return render_template(
        "form.html",
        atividade=atividade,
        titulo=f"Editar Atividade #{atividade_id}",
    )


@app.route("/atividades/<int:atividade_id>/restaurar/<int:versao_id>", methods=["POST"])
@login_required
def restaurar(atividade_id, versao_id):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM versoes WHERE id=? AND atividade_id=?",
            (versao_id, atividade_id),
        ).fetchone()
        if not row:
            flash("Versão não encontrada.", "danger")
            return redirect(url_for("ver", atividade_id=atividade_id))

        snapshot = json.loads(row["snapshot"] or "{}")
        for k in ("id", "criado_em", "atualizado_em"):
            snapshot.pop(k, None)

        # Restaurar gera NOVA versão documentando a reversão (Guia 4.1.6).
        # A versão seguinte deriva da versão ATUAL do registro (não do snapshot).
        cur = conn.execute("SELECT versao FROM atividades WHERE id=?", (atividade_id,)).fetchone()
        versao_atual = (cur["versao"] if cur else None) or "1.0"
        nova_versao = proxima_versao(versao_atual, True)
        sintese = f"v{nova_versao} – reversão à v{row['versao']} (restauração)"
        snapshot["versao"] = nova_versao
        snapshot["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        snapshot["id"] = atividade_id

        _registrar_versao(conn, atividade_id, nova_versao, sintese,
                          _snapshot_json(snapshot), _responsavel_atual())
        set_clause = ",".join(f"{c}=:{c}" for c in _COLUNAS_INSERT)
        conn.execute(
            f"UPDATE atividades SET {set_clause}, atualizado_em=:atualizado_em WHERE id=:id",
            snapshot,
        )

    flash(f"Registro restaurado à v{row['versao']}. Nova versão {nova_versao} criada.", "success")
    return redirect(url_for("ver", atividade_id=atividade_id))


# ── Módulo RIPD ───────────────────────────────────────────────────────────────

_RIPD_COLS_EDITAVEIS = [
    "titulo", "situacao", "versao", "justificativa", "descricao_operacoes",
    "principios", "direitos_titulares", "riscos", "medidas_mitigacao",
    "riscos_residuais", "restricoes_publicacao", "aprovado_por", "aprovado_em",
]


def get_ripd(ripd_id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM ripds WHERE id=?", (ripd_id,)).fetchone()
    return dict(row) if row else None


def _ripd_form_to_dict(form) -> dict:
    principios = {}
    for key, _label in PRINCIPIOS_LGPD:
        v = form.get(f"principios_{key}", "").strip()
        if v:
            principios[key] = v
    direitos = {}
    for key, _label in DIREITOS_TITULARES:
        v = form.get(f"direitos_{key}", "").strip()
        if v:
            direitos[key] = v
    riscos = []
    for linha in (form.get("riscos", "") or "").splitlines():
        linha = linha.strip()
        if not linha:
            continue
        partes = [p.strip() for p in linha.split(";")]
        riscos.append({
            "descricao": partes[0] if partes else "",
            "impacto": partes[1] if len(partes) > 1 else "",
            "probabilidade": partes[2] if len(partes) > 2 else "",
            "aceite": partes[3] if len(partes) > 3 else "",
        })
    return {
        "titulo": form.get("titulo", "").strip(),
        "situacao": form.get("situacao", "rascunho").strip() or "rascunho",
        "justificativa": form.get("justificativa", "").strip(),
        "descricao_operacoes": form.get("descricao_operacoes", "").strip(),
        "principios": json.dumps(principios, ensure_ascii=False),
        "direitos_titulares": json.dumps(direitos, ensure_ascii=False),
        "riscos": json.dumps(riscos, ensure_ascii=False),
        "medidas_mitigacao": json.dumps(parse_lista(form.get("medidas_mitigacao", "")), ensure_ascii=False),
        "riscos_residuais": json.dumps(parse_lista(form.get("riscos_residuais", "")), ensure_ascii=False),
        "restricoes_publicacao": json.dumps(parse_lista(form.get("restricoes_publicacao", "")), ensure_ascii=False),
    }


def _riscos_para_texto(raw) -> str:
    if isinstance(raw, str):
        try:
            raw = json.loads(raw) if raw else []
        except Exception:
            return raw
    if not isinstance(raw, list):
        return ""
    linhas = []
    for r in raw:
        if isinstance(r, dict):
            linhas.append("; ".join(str(r.get(k, "") or "") for k in ("descricao", "impacto", "probabilidade", "aceite")))
        else:
            linhas.append(str(r))
    return "\n".join(linhas)


@app.route("/atividades/<int:atividade_id>/ripd/novo", methods=["POST"])
@login_required
def ripd_novo(atividade_id):
    atividade = get_atividade(atividade_id)
    if not atividade:
        flash("Atividade não encontrada.", "danger")
        return redirect(url_for("listar"))
    with get_conn() as conn:
        exists = conn.execute("SELECT id FROM ripds WHERE atividade_id=? LIMIT 1", (atividade_id,)).fetchone()
        if exists:
            flash("Já existe um RIPD para esta atividade.", "warning")
            return redirect(url_for("ripd_ver", ripd_id=exists["id"]))
        risco = calcular_risco(atividade)
        geral = risco["geral"][0] if risco["geral"] else ""
        espec = risco["especifico"][0] if risco["especifico"] else ""
        just = (
            "Tratamento identificado como de alto risco (Res. CD/ANPD nº 2/2022, art. 4º): "
            f"critério geral — {CRITERIO_GERAL_LABELS.get(geral, geral) or '(não preenchido)'}; "
            f"critério específico — {CRITERIO_ESPECIFICO_LABELS.get(espec, espec) or '(não preenchido)'}."
        )
        cur = conn.execute("""
            INSERT INTO ripds (atividade_id, titulo, situacao, versao, justificativa,
               criterio_geral, criterio_especifico, fatores_risco, alto_risco,
               descricao_operacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (atividade_id, f"RIPD – {atividade['nome_atividade']}", "rascunho", "1.0", just,
              geral, espec, json.dumps(risco["fatores"], ensure_ascii=False),
              1 if risco["alto_risco"] else 0,
              (atividade.get("fluxo_tratamento") or "") + "\n\n" + (atividade.get("finalidade") or "")))
        ripd_id = cur.lastrowid
        row = conn.execute("SELECT * FROM ripds WHERE id=?", (ripd_id,)).fetchone()
        conn.execute("""
            INSERT INTO versoes_ripd (ripd_id, versao, sintese, responsavel, snapshot)
            VALUES (?,?,?,?,?)
        """, (ripd_id, "1.0", "v1.0 – criação do RIPD", _responsavel_atual(), _snapshot_json(dict(row))))
    flash("RIPD criado a partir do gatilho do ROPA.", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


@app.route("/ripds/<int:ripd_id>")
@login_required
def ripd_ver(ripd_id):
    ripd = get_ripd(ripd_id)
    if not ripd:
        flash("RIPD não encontrado.", "danger")
        return redirect(url_for("listar"))
    atividade = get_atividade(ripd["atividade_id"]) or {}
    versoes_ripd = []
    aprovacoes = []
    with get_conn() as conn:
        versoes_ripd = conn.execute(
            "SELECT * FROM versoes_ripd WHERE ripd_id=? ORDER BY id DESC LIMIT 30", (ripd_id,)
        ).fetchall()
        aprovacoes = conn.execute(
            "SELECT * FROM ripd_aprovacoes WHERE ripd_id=? ORDER BY id DESC LIMIT 50", (ripd_id,)
        ).fetchall()
    return render_template("ripd_ver.html", ripd=ripd, atividade=atividade,
                           versoes_ripd=versoes_ripd, aprovacoes=aprovacoes)


@app.route("/ripds/<int:ripd_id>/editar", methods=["GET", "POST"])
@login_required
def ripd_editar(ripd_id):
    ripd = get_ripd(ripd_id)
    if not ripd:
        flash("RIPD não encontrado.", "danger")
        return redirect(url_for("listar"))
    atividade = get_atividade(ripd["atividade_id"]) or {}

    if request.method == "POST":
        novos = _ripd_form_to_dict(request.form)
        with get_conn() as conn:
            estrutural = any(
                novos.get(c) != ripd.get(c) for c in ("justificativa", "riscos", "descricao_operacoes")
            )
            nova_versao = proxima_versao(ripd.get("versao"), estrutural)
            sintese = f"v{nova_versao} – " + _sintese_alteracoes(ripd, novos)
            novos["versao"] = nova_versao
            novos["aprovado_por"] = ripd.get("aprovado_por")
            novos["aprovado_em"] = ripd.get("aprovado_em")
            novos["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            novos["id"] = ripd_id
            _registrar_versao_ripd(conn, ripd_id, nova_versao, sintese,
                                   _snapshot_json({**ripd, **novos}), _responsavel_atual())
            set_clause = ",".join(f"{c}=:{c}" for c in _RIPD_COLS_EDITAVEIS)
            conn.execute(
                f"UPDATE ripds SET {set_clause}, atualizado_em=:atualizado_em WHERE id=:id",
                novos,
            )
        flash(f"RIPD atualizado (versão {nova_versao}).", "success")
        return redirect(url_for("ripd_ver", ripd_id=ripd_id))

    return render_template("ripd_form.html", ripd=ripd, atividade=atividade, titulo=f"Editar RIPD #{ripd_id}")


@app.route("/ripds/<int:ripd_id>/aprovar", methods=["POST"])
@login_required
def ripd_aprovar(ripd_id):
    with get_conn() as conn:
        cur_ripd = conn.execute("SELECT * FROM ripds WHERE id=?", (ripd_id,)).fetchone()
        if not cur_ripd:
            flash("RIPD não encontrado.", "danger")
            return redirect(url_for("listar"))
        ripd = dict(cur_ripd)
        nova_versao = proxima_versao(ripd.get("versao"), True)
        novos = dict(ripd)
        novos["versao"] = nova_versao
        novos["situacao"] = "aprovado"
        novos["aprovado_por"] = _responsavel_atual()
        novos["aprovado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        novos["atualizado_em"] = novos["aprovado_em"]
        sintese = f"v{nova_versao} – aprovação pelo controlador"
        _registrar_versao_ripd(conn, ripd_id, nova_versao, sintese,
                               _snapshot_json({**ripd, **novos}), _responsavel_atual())
        set_clause = ",".join(f"{c}=:{c}" for c in _RIPD_COLS_EDITAVEIS)
        conn.execute(f"UPDATE ripds SET {set_clause}, atualizado_em=:atualizado_em WHERE id=:id", novos)
    flash(f"RIPD aprovado (versão {nova_versao}).", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


@app.route("/ripds/<int:ripd_id>/sugerir-riscos", methods=["POST"])
@login_required
def ripd_sugerir_riscos(ripd_id):
    """Gera riscos típicos a partir dos fatores da atividade (matriz 5×5 ISO 27005).
    Merge com riscos já preenchidos — não duplica ameaças existentes."""
    ripd = get_ripd(ripd_id)
    if not ripd:
        flash("RIPD não encontrado.", "danger")
        return redirect(url_for("listar"))
    atividade = get_atividade(ripd["atividade_id"]) or {}
    atuais = parse_json(ripd.get("riscos")) or []
    if not isinstance(atuais, list):
        atuais = []
    existentes = {r.get("ameaca") for r in atuais if isinstance(r, dict)}
    novos = []
    for r in sugerir_riscos(atividade):
        if r["ameaca"] in existentes:
            continue
        r["aceite"] = "não"  # novo risco ainda não aceite
        novos.append(r)
        existentes.add(r["ameaca"])
    if not novos:
        flash("Nenhum risco novo sugerido — os riscos típicos já estão preenchidos.", "warning")
        return redirect(url_for("ripd_ver", ripd_id=ripd_id))
    combinados = atuais + novos
    with get_conn() as conn:
        nova_versao = proxima_versao(ripd.get("versao"), False)
        sintese = f"v{nova_versao} – sugestão de {len(novos)} risco(s) típico(s) via matriz 5×5"
        novos_estado = {**ripd, "versao": nova_versao, "riscos": json.dumps(combinados, ensure_ascii=False),
                        "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "id": ripd_id}
        _registrar_versao_ripd(conn, ripd_id, nova_versao, sintese,
                               _snapshot_json(novos_estado), _responsavel_atual())
        conn.execute("UPDATE ripds SET riscos=:riscos, versao=:versao, atualizado_em=:atualizado_em WHERE id=:id",
                     novos_estado)
    flash(f"{len(novos)} risco(s) sugerido(s) a partir dos fatores da atividade.", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


@app.route("/ripds/<int:ripd_id>/aprovacoes/solicitar", methods=["POST"])
@login_required
def ripd_aprovacao_solicitar(ripd_id):
    """Abre uma solicitação de aprovação em etapa (papel) para o RIPD."""
    ripd = get_ripd(ripd_id)
    if not ripd:
        flash("RIPD não encontrado.", "danger")
        return redirect(url_for("listar"))
    papel = request.form.get("papel", "").strip()
    if papel not in dict(PAPEIS_APROVACAO):
        flash("Papel de aprovação inválido.", "danger")
        return redirect(url_for("ripd_ver", ripd_id=ripd_id))
    comentario = request.form.get("comentario", "").strip()
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO ripd_aprovacoes (ripd_id, papel, status, comentario) VALUES (?,?,'pendente',?)",
            (ripd_id, papel, comentario),
        )
    flash(f"Aprovação solicitada ao papel: {papel_label(papel)}.", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


@app.route("/ripds/<int:ripd_id>/aprovacoes/<int:aprov_id>/responder", methods=["POST"])
@login_required
def ripd_aprovacao_responder(ripd_id, aprov_id):
    """Registra decisão (aprovar/reprovar) sobre uma solicitação de aprovação."""
    decisao = request.form.get("decisao", "").strip()
    if decisao not in ("aprovar", "reprovar"):
        flash("Decisão inválida.", "danger")
        return redirect(url_for("ripd_ver", ripd_id=ripd_id))
    parecer = request.form.get("parecer", "").strip()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ripd_aprovacoes WHERE id=? AND ripd_id=?", (aprov_id, ripd_id)
        ).fetchone()
        if not row:
            flash("Solicitação de aprovação não encontrada.", "danger")
            return redirect(url_for("ripd_ver", ripd_id=ripd_id))
        conn.execute(
            "UPDATE ripd_aprovacoes SET status=?, aprovador_sub=?, aprovador_nome=?, "
            "comentario=?, respondido_em=datetime('now','localtime') WHERE id=?",
            ("aprovado" if decisao == "aprovar" else "reprovado",
             _sub_atual(), _responsavel_atual(), parecer or row["comentario"], aprov_id),
        )
    flash("Decisão de aprovação registrada.", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


@app.route("/ripds/<int:ripd_id>/aprovacoes/<int:aprov_id>/excluir", methods=["POST"])
@login_required
def ripd_aprovacao_excluir(ripd_id, aprov_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM ripd_aprovacoes WHERE id=? AND ripd_id=?", (aprov_id, ripd_id))
    flash("Solicitação de aprovação removida.", "warning")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))



@app.route("/ripds/<int:ripd_id>/restaurar/<int:versao_id>", methods=["POST"])
@login_required
def ripd_restaurar(ripd_id, versao_id):
    with get_conn() as conn:
        v = conn.execute("SELECT * FROM versoes_ripd WHERE id=? AND ripd_id=?", (versao_id, ripd_id)).fetchone()
        if not v:
            flash("Versão do RIPD não encontrada.", "danger")
            return redirect(url_for("ripd_ver", ripd_id=ripd_id))
        cur = conn.execute("SELECT versao FROM ripds WHERE id=?", (ripd_id,)).fetchone()
        versao_atual = (cur["versao"] if cur else None) or "1.0"
        snapshot = json.loads(v["snapshot"] or "{}")
        for k in ("id", "criado_em", "atualizado_em"):
            snapshot.pop(k, None)
        nova_versao = proxima_versao(versao_atual, True)
        snapshot["versao"] = nova_versao
        snapshot["situacao"] = "rascunho"
        snapshot["aprovado_por"] = None
        snapshot["aprovado_em"] = None
        snapshot["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        snapshot["id"] = ripd_id
        sintese = f"v{nova_versao} – reversão à v{v['versao']} (restauração)"
        _registrar_versao_ripd(conn, ripd_id, nova_versao, sintese,
                               _snapshot_json(snapshot), _responsavel_atual())
        set_clause = ",".join(f"{c}=:{c}" for c in _RIPD_COLS_EDITAVEIS)
        conn.execute(f"UPDATE ripds SET {set_clause}, atualizado_em=:atualizado_em WHERE id=:id", snapshot)
    flash(f"RIPD restaurado à v{v['versao']}. Nova versão {nova_versao} criada.", "success")
    return redirect(url_for("ripd_ver", ripd_id=ripd_id))


def _registrar_versao_ripd(conn, ripd_id, versao, sintese, snapshot, responsavel):
    conn.execute("""
        INSERT INTO versoes_ripd (ripd_id, versao, sintese, responsavel, snapshot)
        VALUES (?,?,?,?,?)
    """, (ripd_id, versao, sintese, responsavel, snapshot))


def _gerar_pdf_ripd(ripd: dict, atividade: dict, publico: bool) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (HRFlowable, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    cor_azul = colors.HexColor("#1F3D7A")
    cor_cinza = colors.HexColor("#5A5A5A")
    cor_leve = colors.HexColor("#EEF2FF")

    titulo_style = ParagraphStyle("tit", fontSize=15, fontName="Helvetica-Bold",
                                  textColor=cor_azul, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle("sub", fontSize=9, fontName="Helvetica",
                               textColor=cor_cinza, alignment=TA_CENTER)
    secao_style = ParagraphStyle("sec", fontSize=11, fontName="Helvetica-Bold",
                                 textColor=cor_azul, spaceBefore=12, spaceAfter=5)
    campo_style = ParagraphStyle("cmp", fontSize=8.5, fontName="Helvetica", leading=12)
    label_style = ParagraphStyle("lbl", fontSize=8.5, fontName="Helvetica-Bold", textColor=cor_cinza)
    rodape_style = ParagraphStyle("rod", fontSize=7, fontName="Helvetica", textColor=cor_cinza, alignment=TA_CENTER)

    story = []
    story.append(Paragraph("RELATÓRIO DE IMPACTO À PROTEÇÃO DE DADOS PESSOAIS (RIPD)", titulo_style))
    story.append(Paragraph("PPSI 2.0 – controle 23.3 · LGPD, art. 38", sub_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=cor_azul))
    if publico:
        story.append(Paragraph("VERSÃO PUBLICIZÁVEL", ParagraphStyle(
            "pub", fontSize=9, fontName="Helvetica-Bold", textColor=colors.red, spaceBefore=6)))
    story.append(Spacer(1, 0.3*cm))

    def sec(titulo):
        story.append(Paragraph(titulo, secao_style))

    def linha(label, val):
        v = str(val or "").strip() or "—"
        tbl = Table([[Paragraph(label, label_style), Paragraph(v, campo_style)]],
                    colWidths=[4.5*cm, 12*cm])
        tbl.setStyle(TableStyle([
            ("FONTSIZE", (0,0), (-1,-1), 8.5),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, cor_leve]),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("BOX", (0,0), (-1,-1), 0.4, cor_cinza),
            ("INNERGRID", (0,0), (-1,-1), 0.3, cor_cinza),
        ]))
        story.append(tbl)

    # 1. Contexto
    sec("1. Contexto do produto/serviço")
    linha("Atividade (ROPA)", f"#{atividade.get('id','')} {atividade.get('nome_atividade','')}")
    linha("Título do RIPD", ripd.get("titulo"))
    linha("Unidade responsável", atividade.get("unidade_controladora"))
    linha("Situação", ripd.get("situacao"))
    linha("Versão", ripd.get("versao"))
    linha("Justificativa (gatilho)", ripd.get("justificativa"))
    fatores = parse_json(ripd.get("fatores_risco"))
    if fatores:
        linha("Fatores de risco identificados", "; ".join(fatores))

    # 2. Tratamento
    sec("2. Tratamento dos dados pessoais")
    linha("Descrição das operações", ripd.get("descricao_operacoes"))
    linha("Base legal", f"{atividade.get('base_legal','')} – {BASES_LEGAIS.get(atividade.get('base_legal',''),'')}")
    linha("Titulares", atividade.get("categorias_titulares"))
    linha("Tipos de dados", atividade.get("categorias_dados"))
    linha("Fluxo", atividade.get("fluxo_tratamento"))
    linha("Compartilhamento", atividade.get("destinatarios"))
    linha("Transferência internacional", lista_para_texto(parse_json(ripd.get("transferencia_inter")) or atividade.get("transferencia_inter")))

    # 3. Princípios
    sec("3. Análise dos princípios da LGPD (art. 6º)")
    principios = parse_json(ripd.get("principios"))
    if principios:
        for key, label in PRINCIPIOS_LGPD:
            if principios.get(key):
                linha(label, principios[key])
    else:
        linha("Princípios", "Não preenchido")

    # 4. Direitos
    sec("4. Garantia dos direitos dos titulares (art. 18)")
    direitos = parse_json(ripd.get("direitos_titulares"))
    if direitos:
        for key, label in DIREITOS_TITULARES:
            if direitos.get(key):
                linha(label, direitos[key])
    else:
        linha("Direitos", "Não preenchido")

    # 5. Gestão de riscos (RESTRITO na versão publicizável)
    sec("5. Gestão de riscos")
    if publico:
        story.append(Paragraph(
            "<b>Conteúdo restrito.</b> A descrição de riscos, medidas de mitigação e riscos residuais foi "
            "suprimida nesta versão por tratar-se de informação crítica de segurança da informação "
            "(Portaria SGD/MGI nº 9.511/2025, art. 2º, III e art. 5º).", campo_style))
        restr = parse_json(ripd.get("restricoes_publicacao"))
        if restr:
            story.append(Spacer(1, 0.2*cm))
            linha("Conteúdo suprimido", "; ".join(restr))
    else:
        riscos = parse_json(ripd.get("riscos"))
        if riscos:
            for i, r in enumerate(riscos, 1):
                if isinstance(r, dict):
                    linha(f"Risco {i}", f"Descrição: {r.get('descricao','')} · Impacto: {r.get('impacto','')} · "
                          f"Probabilidade: {r.get('probabilidade','')} · Aceite: {r.get('aceite','')}")
                else:
                    linha(f"Risco {i}", str(r))
        else:
            linha("Riscos", "Não preenchido")
        linha("Medidas de mitigação", "; ".join(parse_json(ripd.get("medidas_mitigacao"))))
        linha("Riscos residuais", "; ".join(parse_json(ripd.get("riscos_residuais"))))

    # 6. Aprovação
    sec("6. Aprovação")
    linha("Aprovação (controlador)", ripd.get("aprovado_por") or "—")
    linha("Data da aprovação", ripd.get("aprovado_em") or "—")

    story.append(Spacer(1, 0.8*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))
    story.append(Paragraph(
        f"{ORGANIZACAO} · {UNIDADE} · Encarregado: {ENCARREGADO} · "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", rodape_style))

    doc.build(story)
    buf.seek(0)
    return buf.read()


@app.route("/ripds/<int:ripd_id>/exportar")
@login_required
def ripd_exportar(ripd_id):
    ripd = get_ripd(ripd_id)
    if not ripd:
        flash("RIPD não encontrado.", "danger")
        return redirect(url_for("listar"))
    atividade = get_atividade(ripd["atividade_id"]) or {}
    publico = request.args.get("publico", "0") == "1"
    pdf_bytes = _gerar_pdf_ripd(ripd, atividade, publico)
    nome = f"ripd_{ripd_id}{'_publico' if publico else ''}.pdf"
    return Response(
        pdf_bytes, mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nome}"},
    )


@app.route("/atividades/<int:atividade_id>/excluir", methods=["POST"])
@login_required
def excluir(atividade_id):
    with get_conn() as conn:
        conn.execute("UPDATE atividades SET ativo=0 WHERE id=?", (atividade_id,))
    flash(f"Atividade #{atividade_id} removida.", "warning")
    return redirect(url_for("listar"))


@app.route("/atividades/<int:atividade_id>/reabrir", methods=["POST"])
@login_required
def reabrir(atividade_id):
    """Reabre uma atividade previamente descartada (soft-delete reversível)."""
    with get_conn() as conn:
        conn.execute("UPDATE atividades SET ativo=1 WHERE id=?", (atividade_id,))
    flash(f"Atividade #{atividade_id} reaberta.", "success")
    return redirect(url_for("ver", atividade_id=atividade_id))



@app.route("/validar")
@login_required
def validar():
    registros = todos_registros()
    resultados = []
    for r in registros:
        score, faltando = pontuacao(r)
        resultados.append({"atividade": r, "score": score, "faltando": faltando})
    scores = [x["score"] for x in resultados]
    media = sum(scores) / len(scores) if scores else 0
    return render_template("validar.html", resultados=resultados, media=media)


@app.route("/importar", methods=["GET", "POST"])
@login_required
def importar():
    if request.method == "GET":
        return render_template("importar.html")

    # POST: Handle file upload
    if "file" not in request.files:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("importar"))

    file = request.files["file"]
    if file.filename == "":
        flash("Arquivo vazio.", "danger")
        return redirect(url_for("importar"))

    # Validate file type
    if not file.filename.endswith(".json"):
        flash("Apenas arquivos .json são aceitos.", "danger")
        return redirect(url_for("importar"))

    # Validate file size (max 10MB)
    file.seek(0, 2)  # Seek to end
    size = file.tell()
    file.seek(0)  # Reset to start
    if size > 10 * 1024 * 1024:
        flash("Arquivo muito grande (máximo 10MB).", "danger")
        return redirect(url_for("importar"))

    # Get conflict resolution strategy
    strategy = request.form.get("conflict_strategy", "skip")
    if strategy not in ("skip", "merge", "overwrite"):
        strategy = "skip"

    try:
        # Read and import
        content = file.read().decode("utf-8")
        imported, skipped, errors, error_msgs = cnil_pia_importer.import_from_content(
            content=content,
            conflict_strategy=strategy
        )

        if errors > 0:
            flash(
                f"Importação com problemas: {imported} inseridos, {skipped} pulados, {errors} erros.",
                "warning"
            )
            for msg in error_msgs[:5]:  # Show first 5 errors
                flash(f"  • {msg}", "info")
        else:
            flash(
                f"✓ Importação bem-sucedida: {imported} atividades importadas, {skipped} puladas.",
                "success"
            )

        return redirect(url_for("listar"))

    except Exception as e:
        flash(f"Erro ao importar: {str(e)}", "danger")
        return redirect(url_for("importar"))


@app.route("/exportar")
@login_required
def exportar():
    fmt = request.args.get("formato", "json")
    registros = todos_registros()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if fmt == "json":
        data = json.dumps(registros, ensure_ascii=False, indent=2, default=str)
        return Response(
            data,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.json"},
        )

    if fmt == "csv":
        output = io.StringIO()
        if registros:
            writer = csv.DictWriter(output, fieldnames=registros[0].keys())
            writer.writeheader()
            writer.writerows(registros)
        return Response(
            "\ufeff" + output.getvalue(),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.csv"},
        )

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            flash("openpyxl não instalado.", "danger")
            return redirect(url_for("index"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RoPA"

        HEADER_FILL = PatternFill("solid", fgColor="1F3D7A")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        SENSIVEL_FILL = PatternFill("solid", fgColor="FFE0E0")
        ALT_FILL = PatternFill("solid", fgColor="F0F4FF")

        colunas = [
            ("id", "ID", 6),
            ("nome_atividade", "Atividade", 36),
            ("situacao", "Situação", 12),
            ("versao", "Versão", 8),
            ("finalidade", "Finalidade", 30),
            ("base_legal", "Base Legal", 8),
            ("unidade_controladora", "Unidade", 20),
            ("responsavel_preenchimento", "Responsável", 22),
            ("previsao_normativa", "Previsão normativa", 22),
            ("categorias_titulares", "Titulares", 22),
            ("titulares_estimativa", "Estimativa titulares", 20),
            ("titulares_protecao_reforcada", "Proteção reforçada", 18),
            ("categorias_dados", "Dados", 24),
            ("tipos_dados", "Tipos de dados (FCI)", 30),
            ("dados_sensiveis", "Sensível", 9),
            ("tipos_dados_sensiveis", "Tipos sensíveis", 22),
            ("fluxo_tratamento", "Fluxo", 30),
            ("origem_dados", "Origem", 22),
            ("local_armazenamento", "Armazenamento", 22),
            ("prazo_retencao", "Retenção", 20),
            ("eliminacao_destinacao", "Eliminação", 22),
            ("frequencia_tratamento", "Frequência", 14),
            ("controladores", "Controladores", 22),
            ("operadores", "Operadores", 22),
            ("destinatarios", "Compartilhamento", 22),
            ("compartilhamentos", "Compartilhamentos", 24),
            ("transferencia_inter", "Transf. int.", 14),
            ("transferencia_internacional", "Transf. int. detalhada", 24),
            ("medidas_seguranca", "Medidas Seg.", 30),
            ("sistema_sei", "SEI", 18),
            ("observacoes", "Observações", 22),
            ("atualizado_em", "Atualizado", 16),
        ]

        def _render(campo, val):
            if campo == "dados_sensiveis":
                return "Sim" if val else "Não"
            if campo == "base_legal" and val:
                return f"{val} – {BASES_LEGAIS.get(val, val)}"
            if campo == "titulares_estimativa":
                return dict_estimativa_para_texto(val).replace("\n", "; ")
            if campo == "tipos_dados":
                return dict_tipos_para_texto(val).replace("\n", "; ")
            if campo in JSON_FIELDS:
                return lista_para_texto(val).replace("\n", "; ")
            return str(val or "")

        for col_idx, (_, titulo, largura) in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura
        ws.row_dimensions[1].height = 30

        for row_idx, reg in enumerate(registros, 2):
            fill = SENSIVEL_FILL if reg.get("dados_sensiveis") else (ALT_FILL if row_idx % 2 == 0 else None)
            for col_idx, (campo, _, _) in enumerate(colunas, 1):
                val = _render(campo, reg.get(campo, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.xlsx"},
        )

    if fmt == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (HRFlowable, PageBreak, Paragraph,
                                            SimpleDocTemplate, Spacer, Table,
                                            TableStyle)
        except ImportError:
            flash("reportlab não instalado.", "danger")
            return redirect(url_for("index"))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=2.5 * cm, bottomMargin=2 * cm,
        )

        cor_azul = colors.HexColor("#1F3D7A")
        cor_cinza = colors.HexColor("#5A5A5A")
        cor_leve = colors.HexColor("#EEF2FF")

        titulo_style = ParagraphStyle("titulo", fontSize=16, fontName="Helvetica-Bold",
                                      textColor=cor_azul, spaceAfter=4, alignment=TA_CENTER)
        sub_style = ParagraphStyle("sub", fontSize=10, fontName="Helvetica",
                                   textColor=cor_cinza, spaceAfter=2, alignment=TA_CENTER)
        secao_style = ParagraphStyle("secao", fontSize=11, fontName="Helvetica-Bold",
                                     textColor=cor_azul, spaceBefore=12, spaceAfter=6)
        campo_style = ParagraphStyle("campo", fontSize=8, fontName="Helvetica",
                                     textColor=colors.black, leading=11)
        label_style = ParagraphStyle("label", fontSize=8, fontName="Helvetica-Bold",
                                     textColor=cor_cinza, leading=11)
        rodape_style = ParagraphStyle("rodape", fontSize=7, fontName="Helvetica",
                                      textColor=cor_cinza, alignment=TA_CENTER)

        story = []
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph(ORGANIZACAO, titulo_style))
        story.append(Paragraph(UNIDADE, sub_style))
        story.append(Spacer(1, 0.3 * cm))
        story.append(HRFlowable(width="100%", thickness=2, color=cor_azul))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph("REGISTRO DE ATIVIDADES DE TRATAMENTO", titulo_style))
        story.append(Paragraph("Conforme LGPD – Lei 13.709/2018, Art. 37", sub_style))
        story.append(Spacer(1, 0.5 * cm))

        data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")
        meta = [
            ["Data de geração:", data_geracao, "Encarregado (DPO):", ENCARREGADO],
            ["Total de atividades:", str(len(registros)), "Versão:", f"RoPA-{datetime.now().strftime('%Y%m%d')}"],
        ]
        t_meta = Table(meta, colWidths=[4 * cm, 6 * cm, 4 * cm, 5.5 * cm])
        t_meta.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [cor_leve, colors.white]),
            ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

        story.append(Paragraph("Resumo de Completude", secao_style))
        scores_info = [(r["id"], r["nome_atividade"], pontuacao(r)[0]) for r in registros]
        sum_data = [["ID", "Atividade de Tratamento", "Completude", "Status"]]
        for rid, nome, score in scores_info:
            status = "Completo" if score >= 80 else ("Parcial" if score >= 50 else "Incompleto")
            cor_st = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)
            sum_data.append([str(rid), nome[:55], f"{score}%",
                             Paragraph(f'<font color="{cor_st.hexval()}">{status}</font>', campo_style)])
        t_sum = Table(sum_data, colWidths=[1.2 * cm, 10 * cm, 2.5 * cm, 3 * cm])
        t_sum.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), cor_azul),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, cor_leve]),
            ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t_sum)

        media_geral = sum(s for _, _, s in scores_info) / len(scores_info) if scores_info else 0
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(
            f"<b>Média de completude: {media_geral:.1f}%</b>  ·  "
            f"{sum(1 for _, _, s in scores_info if s >= 80)} registro(s) completo(s) de {len(registros)}",
            ParagraphStyle("media", fontSize=8, fontName="Helvetica", textColor=cor_cinza),
        ))

        story.append(PageBreak())
        story.append(Paragraph("Fichas das Atividades de Tratamento", secao_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

        campos_ficha = [
            ("situacao", "Situação do registro"),
            ("versao", "Versão"),
            ("responsavel_preenchimento", "Responsável pelo preenchimento"),
            ("unidade_controladora", "Unidade controladora"),
            ("sistema_sei", "Processo SEI relacionado"),
            ("finalidade", "Finalidade"),
            ("base_legal", "Base legal (LGPD)"),
            ("previsao_normativa", "Previsão normativa específica"),
            ("categorias_titulares", "Categorias de titulares"),
            ("titulares_estimativa", "Estimativa de titulares"),
            ("titulares_protecao_reforcada", "Titulares com proteção reforçada"),
            ("categorias_dados", "Dados pessoais envolvidos"),
            ("tipos_dados", "Tipos de dados (FCI-ANPD)"),
            ("dados_sensiveis", "Dados sensíveis (Art. 5º, II)"),
            ("tipos_dados_sensiveis", "Tipos de dados sensíveis"),
            ("fluxo_tratamento", "Fluxo de tratamento"),
            ("origem_dados", "Origem dos dados"),
            ("local_armazenamento", "Local e meio de armazenamento"),
            ("prazo_retencao", "Prazo de retenção"),
            ("eliminacao_destinacao", "Eliminação/destinação final"),
            ("frequencia_tratamento", "Frequência do tratamento"),
            ("controladores", "Controladores"),
            ("operadores", "Operadores"),
            ("destinatarios", "Compartilhamento"),
            ("compartilhamentos", "Compartilhamentos detalhados"),
            ("transferencia_inter", "Transferência internacional"),
            ("transferencia_internacional", "Transferência internacional detalhada"),
            ("medidas_seguranca", "Medidas de segurança (Art. 46)"),
            ("observacoes", "Observações"),
            ("criado_em", "Data de criação"),
            ("atualizado_em", "Última atualização"),
        ]

        def _render_pdf(campo, val):
            if campo == "dados_sensiveis":
                return "Sim" if val else "Não"
            if campo == "base_legal" and val:
                return f"{val} – {BASES_LEGAIS.get(val, val)}"
            if campo == "titulares_estimativa":
                return dict_estimativa_para_texto(val)
            if campo == "tipos_dados":
                return dict_tipos_para_texto(val)
            if campo in JSON_FIELDS:
                return lista_para_texto(val)
            return str(val or "")

        for reg in registros:
            story.append(Spacer(1, 0.4 * cm))
            score, _ = pontuacao(reg)
            cor_score = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)
            cabecalho = Table(
                [[Paragraph(f"<b>#{reg['id']}  {reg['nome_atividade']}</b>",
                            ParagraphStyle("cabe", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white)),
                  Paragraph(f'<font color="{cor_score.hexval()}"><b>{score}%</b></font>',
                            ParagraphStyle("pct", fontSize=10, fontName="Helvetica-Bold",
                                           textColor=colors.white, alignment=TA_CENTER))]],
                colWidths=[14 * cm, 2.5 * cm],
            )
            cabecalho.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), cor_azul),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (0, -1), 8),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(cabecalho)

            ficha_data = []
            for campo, label in campos_ficha:
                val = _render_pdf(campo, reg.get(campo, ""))
                val_str = str(val).strip() if val else "—"
                ficha_data.append([Paragraph(label, label_style), Paragraph(val_str, campo_style)])

            t_ficha = Table(ficha_data, colWidths=[5 * cm, 11.5 * cm])
            t_ficha.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, cor_leve]),
                ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(t_ficha)

        story.append(Spacer(1, 1 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"{ORGANIZACAO}  ·  {UNIDADE}  ·  "
            f"Encarregado: {ENCARREGADO}  ·  Gerado em {data_geracao}",
            rodape_style,
        ))
        story.append(Paragraph(
            NORMAS_RODAPE,
            rodape_style,
        ))
        doc.build(story)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=relatorio_ropa_{ts}.pdf"},
        )

    flash("Formato inválido.", "danger")
    return redirect(url_for("index"))


@app.route("/seed", methods=["POST"])
@login_required
def seed():
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM atividades WHERE ativo=1").fetchone()[0]
        if count > 0:
            flash("Base já possui registros. Seed não executado.", "warning")
            return redirect(url_for("index"))
        for ex in EXEMPLOS:
            ex["situacao"] = "concluido"
            ex["versao"] = "1.0"
            conn.execute("""
                INSERT INTO atividades
                  (nome_atividade, finalidade, base_legal, categorias_titulares,
                   categorias_dados, dados_sensiveis, destinatarios, transferencia_inter,
                   prazo_retencao, medidas_seguranca, unidade_controladora, sistema_sei,
                   observacoes, situacao, versao)
                VALUES
                  (:nome_atividade,:finalidade,:base_legal,:categorias_titulares,
                   :categorias_dados,:dados_sensiveis,:destinatarios,:transferencia_inter,
                   :prazo_retencao,:medidas_seguranca,:unidade_controladora,:sistema_sei,
                   :observacoes,:situacao,:versao)
            """, ex)
    flash(f"{len(EXEMPLOS)} atividades de exemplo inseridas.", "success")
    return redirect(url_for("index"))


# ── Helper ────────────────────────────────────────────────────────────────────

def _form_to_dict(form) -> dict:
    def _j_lista(texto):
        return json.dumps(parse_lista(texto), ensure_ascii=False)

    def _j_dict(texto):
        return json.dumps(parse_dict_tipos(texto), ensure_ascii=False)

    def _j_estimativa(texto):
        return json.dumps(parse_estimativa(texto), ensure_ascii=False)

    prot_ref = form.getlist("titulares_protecao_reforcada")
    transf_int = parse_lista(form.get("transferencia_internacional", ""))
    if not transf_int:
        transf_int = ["N/A"]

    return dict(
        nome_atividade=form.get("nome_atividade", "").strip(),
        finalidade=form.get("finalidade", "").strip(),
        base_legal=form.get("base_legal", "").strip(),
        categorias_titulares=form.get("categorias_titulares", "").strip(),
        categorias_dados=form.get("categorias_dados", "").strip(),
        dados_sensiveis=1 if form.get("dados_sensiveis") else 0,
        destinatarios=form.get("destinatarios", "").strip(),
        transferencia_inter=form.get("transferencia_inter", "N/A").strip(),
        prazo_retencao=form.get("prazo_retencao", "").strip(),
        medidas_seguranca=form.get("medidas_seguranca", "").strip(),
        unidade_controladora=form.get("unidade_controladora", "").strip(),
        sistema_sei=form.get("sistema_sei", "").strip(),
        observacoes=form.get("observacoes", "").strip(),
        responsavel_preenchimento=form.get("responsavel_preenchimento", "").strip(),
        situacao=form.get("situacao", "").strip() or "em_andamento",
        previsao_normativa=form.get("previsao_normativa", "").strip(),
        titulares_estimativa=_j_estimativa(form.get("titulares_estimativa", "")),
        titulares_protecao_reforcada=json.dumps(prot_ref, ensure_ascii=False),
        tipos_dados=_j_dict(form.get("tipos_dados", "")),
        tipos_dados_sensiveis=_j_lista(form.get("tipos_dados_sensiveis", "")),
        fluxo_tratamento=form.get("fluxo_tratamento", "").strip(),
        origem_dados=_j_lista(form.get("origem_dados", "")),
        local_armazenamento=form.get("local_armazenamento", "").strip(),
        eliminacao_destinacao=form.get("eliminacao_destinacao", "").strip(),
        frequencia_tratamento=form.get("frequencia_tratamento", "").strip(),
        controladores=_j_lista(form.get("controladores", "")),
        operadores=_j_lista(form.get("operadores", "")),
        compartilhamentos=_j_lista(form.get("compartilhamentos", "")),
        transferencia_internacional=json.dumps(transf_int, ensure_ascii=False),
        tecnologias_emergentes=1 if form.get("tecnologias_emergentes") else 0,
        decisoes_automatizadas=1 if form.get("decisoes_automatizadas") else 0,
        vigilancia_zonas_publicas=1 if form.get("vigilancia_zonas_publicas") else 0,
    )


_COLUNAS_INSERT = [
    "nome_atividade", "finalidade", "base_legal", "categorias_titulares",
    "categorias_dados", "dados_sensiveis", "destinatarios", "transferencia_inter",
    "prazo_retencao", "medidas_seguranca", "unidade_controladora", "sistema_sei",
    "observacoes", "responsavel_preenchimento", "situacao", "versao",
    "titulares_estimativa", "titulares_protecao_reforcada", "tipos_dados",
    "tipos_dados_sensiveis", "fluxo_tratamento", "origem_dados",
    "local_armazenamento", "eliminacao_destinacao", "frequencia_tratamento",
    "previsao_normativa", "controladores", "operadores", "compartilhamentos",
    "transferencia_internacional", "tecnologias_emergentes", "decisoes_automatizadas",
    "vigilancia_zonas_publicas",
]


def _snapshot_json(atividade: dict) -> str:
    return json.dumps(atividade, ensure_ascii=False, default=str)


def _registrar_versao(conn, atividade_id: int, versao: str, sintese: str,
                      snapshot: str, responsavel: str):
    conn.execute("""
        INSERT INTO versoes (atividade_id, versao, sintese, responsavel, snapshot)
        VALUES (?,?,?,?,?)
    """, (atividade_id, versao, sintese, responsavel, snapshot))


def _sintese_alteracoes(atual: dict, novos: dict) -> str:
    alterados = [desc for campo, (desc, _) in CAMPOS_VALIDACAO.items()
                 if novos.get(campo) != atual.get(campo)]
    return ", ".join(alterados[:6]) if alterados else "ajuste menor"


def _responsavel_atual() -> str:
    u = session.get("user", {}) or {}
    return u.get("name") or u.get("username") or "Web"


def _sub_atual() -> str:
    u = session.get("user", {}) or {}
    return str(u.get("sub") or u.get("id") or u.get("username") or "anon")



# ── Entry point ───────────────────────────────────────────────────────────────

# Inicializa DB no import (cobre gunicorn e `python app.py`)
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
